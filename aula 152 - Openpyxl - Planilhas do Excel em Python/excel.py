"""
Documentação:
https://openpyxl.readthedocs.io/en/stable/

pip install openpyxl
"""
import openpyxl
from random import uniform

pedidos = openpyxl.load_workbook('pedidos.xlsx')
nome_panilhas = pedidos.sheetnames
planilha1 = pedidos['Página1']

for linha in range(5, 10):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

pedidos.save('nova_planilha.xlsx')
